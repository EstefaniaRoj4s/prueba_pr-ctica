import pandas as pd
import openpyxl
import datetime
from collections import defaultdict
import sys


def load_data(path):
    wb = openpyxl.load_workbook(path, data_only=True)
    productos = {}
    ws = wb['productos']
    for rows in ws.iter_rows(min_row=2, values_only=True):
        producto_id, nombre, precio_unitario, iva, stock_actual = rows
        if producto_id is None:
            continue
        productos[producto_id] = {
            'nombre': nombre or "",
            'precio_unitario': float(precio_unitario) if precio_unitario is not None else 0.0,
            'iva': float(iva) if iva is not None else 0.0,
            'stock_actual': int(stock_actual) if stock_actual is not None else 0
        }
        

    pedidos = {}
    ws = wb ['pedidos']
    for rows in ws.iter_rows(min_row=2, values_only=True):
        pedido_id, cliente, fecha = rows
        if pedido_id is None:
            continue
        pedidos[pedido_id] = {
            'cliente_id': cliente or "",
            'fecha_pedido': fecha if isinstance(fecha, datetime.date) else None
        }

    lineas = []
    ws = wb ['lineas_pedido']
    for rows in ws.iter_rows(min_row=2, values_only=True):
        pedido_id, producto_id, cantidad, descuento = rows
        if pedido_id is None or producto_id is None:
            continue
        lineas.append({
            'pedido_id': pedido_id,
            'producto_id': producto_id,
            'cantidad': int(cantidad) if cantidad is not None else 0,
            'descuento': float(descuento) if descuento is not None else 0.0
        })

    facturas = {}
    ws = wb['facturas']
    for rows in ws.iter_rows(min_row=2, values_only=True):
        factura_id, cliente, vencimiento, monto, pagado, estado= rows
        if factura_id is None:
            continue
        facturas[factura_id] = {
            'factura_id': factura_id,
            'cliente': cliente or "",
            'vencimiento': vencimiento if isinstance(vencimiento, datetime.datetime) else None,
            'monto': float(monto) if monto is not None else 0.0,
            'pagado': float(pagado) if pagado is not None else 0.0,
            'estado': estado or ""  
        }
    return productos, pedidos, lineas, facturas

def reporte_total_pedido(productos, pedidos, lineas):
    resultados =[]
    pedido_lineas = defaultdict(list)
    for linea in lineas:
        pedido_lineas[linea['pedido_id']].append(linea)
    for pid, lin in pedido_lineas.items():
        base_imponible = 0.0
        iva_total = 0.0
        for linea in lin:
            producto = productos.get(linea['producto_id'])
            if producto is None:
                continue
            precio_unitario = producto['precio_unitario']
            iva = producto['iva']
            cantidad = linea['cantidad']
            descuento = linea['descuento']
            subtotal = precio_unitario * cantidad * (1 - descuento / 100)
            base_imponible += subtotal
            iva_total += subtotal * iva / 100

        base_imponible = round(base_imponible, 2)
        iva_total = round(iva_total, 2)
        total_pedido = round(base_imponible + iva_total, 2)
        cliente = pedidos.get(pid, {}).get('cliente_id', 'Desconocido')
        resultados.append({
            'pedido_id': pid,
            'cliente': cliente,
            'base_imponible': base_imponible,
            'iva_total': iva_total,
            'total_pedido': total_pedido
        })
    return resultados

def reporte_quiebre_stock(productos, lineas):
    demanda = defaultdict(int)
    for li in lineas:
        if li["cantidad"] > 0:
            demanda[li["producto_id"]] += li["cantidad"]

    resultados = []
    for pid, prod in productos.items():      
        dem = demanda.get(pid, 0)           
        stock = prod["stock_actual"]
        if dem > stock:
            resultados.append((pid, prod["nombre"], stock, dem, dem - stock))

    resultados.sort(key=lambda x: x[4], reverse=True)
    return resultados

def main():
    
    path = 'Dsetprueba.xlsx'
    
    productos, pedidos, lineas, facturas = load_data(path)
    reporte1 = reporte_total_pedido(productos, pedidos, lineas)
    reporte2 = reporte_quiebre_stock(productos, lineas)
    print("Reporte Total Pedido:")
    for r in reporte1:
        print(r)
    print("\nReporte Quiebre de Stock:")
    for r in reporte2:
        print(r)

if __name__ == "__main__":    
    main()


